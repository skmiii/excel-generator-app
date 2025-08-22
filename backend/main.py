from fastapi import FastAPI
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from typing import List, Optional, Dict
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
import io

app = FastAPI()

# CORS設定
origins = [
    "http://localhost:3000",
    "http://localhost:5173", # Viteのデフォルトポート
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- データモデルの定義 ---
class DropdownColumn(BaseModel):
    name: str
    options: List[str]

class CustomColumn(BaseModel):
    name: str
    type: str # 'free' or 'dropdown'
    options: Optional[List[str]] = None

class ExcelRequest(BaseModel):
    dynamic_columns: List[str] = Field(default_factory=list)
    custom_columns: List[CustomColumn] = Field(default_factory=list)

# --- ダミーデータ ---
# 本来はデータベースで管理する
industries = ["IT", "製造", "金融", "不動産", "サービス"]
prefectures = [
    "北海道", "青森県", "岩手県", "宮城県", "秋田県", "山形県", "福島県", "茨城県", "栃木県", "群馬県",
    "埼玉県", "千葉県", "東京都", "神奈川県", "新潟県", "富山県", "石川県", "福井県", "山梨県", "長野県",
    "岐阜県", "静岡県", "愛知県", "三重県", "滋賀県", "京都府", "大阪府", "兵庫県", "奈良県", "和歌山県",
    "鳥取県", "島根県", "岡山県", "広島県", "山口県", "徳島県", "香川県", "愛媛県", "高知県", "福岡県",
    "佐賀県", "長崎県", "熊本県", "大分県", "宮崎県", "鹿児島県", "沖縄県"
]


# --- APIエンドポイント ---

@app.get("/api/industries")
async def get_industries():
    """業種の一覧を取得する"""
    return industries

@app.post("/api/generate-excel")
async def generate_excel(request: ExcelRequest):
    """Excelファイルを生成する"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "顧客リスト"

    # --- ヘッダーの定義 ---
    # 必須カラム
    headers = [
        "顧客法人名",
        "担当者名（姓）",
        "担当者名（名）",
        "電話番号",
        "業種",
    ]
    
    # 自由追加カラム
    dynamic_header_map = {
        "prefecture": "県域",
        "address": "住所",
        "email": "メールアドレス",
        "inflow_date": "流入日",
        "inflow_source": "流入元",
        "list_name": "リスト名",
    }
    for col_key in request.dynamic_columns:
        headers.append(dynamic_header_map.get(col_key, col_key))

    # カスタムカラム
    for col in request.custom_columns:
        headers.append(col.name)

    ws.append(headers)

    # --- データバリデーション（プルダウン）の設定 ---
    # 業種
    industry_col_index = headers.index("業種") + 1
    dv_industry = DataValidation(type="list", formula1=f'"{",".join(industries)}"')
    dv_industry.add(f'{openpyxl.utils.get_column_letter(industry_col_index)}2:{openpyxl.utils.get_column_letter(industry_col_index)}1048576')
    ws.add_data_validation(dv_industry)

    # 県域
    if "県域" in headers:
        pref_col_index = headers.index("県域") + 1
        dv_pref = DataValidation(type="list", formula1=f'"{",".join(prefectures)}"')
        dv_pref.add(f'{openpyxl.utils.get_column_letter(pref_col_index)}2:{openpyxl.utils.get_column_letter(pref_col_index)}1048576')
        ws.add_data_validation(dv_pref)
        
    # カスタムカラムのプルダウン
    for col in request.custom_columns:
        if col.type == 'dropdown' and col.options:
            col_index = headers.index(col.name) + 1
            dv_custom = DataValidation(type="list", formula1=f'"{",".join(col.options)}"')
            dv_custom.add(f'{openpyxl.utils.get_column_letter(col_index)}2:{openpyxl.utils.get_column_letter(col_index)}1048576')
            ws.add_data_validation(dv_custom)


    # メモリ上のファイルに保存
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=customer_list_format.xlsx"}
    )

@app.get("/")
def read_root():
    return {"message": "Excel Generator API is running!"}
